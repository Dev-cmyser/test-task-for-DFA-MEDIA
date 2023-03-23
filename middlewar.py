from sqlalchemy.orm import Session
from models import engine
from models import Company, Qoil, Qliq, Forecast, Fact
from openpyxl import load_workbook


def add_data_to_db(data):
    with Session(autoflush=False, bind=engine) as db:
        # print(list(data.items()))
        if list(data.items()) == []:
            return
        try:
            key, arr = list(data.items())[0]
        except ValueError or IndexError:
            return
        company = Company(name=key)
        db.add(company)  
        db.commit() 
        fact = Fact(company_id=company.id)
        forecast = Forecast(company_id=company.id)
        for item in [fact, forecast]:
            db.add(item)  
        db.commit() 
        print(company.id, fact.id, forecast.id)
        
        # fact Qliq
        qliq_fact = Qliq(data_1=arr[0], data_2=arr[1], fact_id=fact.id)
        qoil_fact = Qoil(data_1=arr[0], data_2=arr[1], fact_id=fact.id)

        qliq_forecast = Qliq(data_1=arr[0], data_2=arr[1], forecast_id=forecast.id)
        qoil_forecast = Qoil(data_1=arr[0], data_2=arr[1], forecast_id=forecast.id)

        for item in [qliq_fact, qoil_fact, qliq_forecast, qoil_forecast]:
            db.add(item)  
        db.commit() 


def parse_exel() -> dict:
    wb = load_workbook("exel.xlsx")
    ws = wb.active
    count = 0
    companies = {}
    for row in ws.iter_rows():
        # company = ''
        for cell in row:
            if (
                cell.value == 0
                or cell.value == None
                or cell.value
                in ["company", "fact", "forecast", "Qliq", "Qoil", "data1", "data2"]
            ):
                continue
            if cell.value == "company1" or cell.value == "company2":
                companies[cell.value] = []
                company = cell.value
            else:
                companies[company].append(cell.value)
        yield companies
        companies = {}
